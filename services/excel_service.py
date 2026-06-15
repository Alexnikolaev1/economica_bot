"""Генерация Excel-отчётов по операциям."""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from database import FinancialSummary

HEADERS = [
    "ID",
    "Дата",
    "Тип",
    "Сумма",
    "Валюта",
    "Сумма (база)",
    "Категория",
    "Контрагент",
    "Описание",
    "Создано",
]

TYPE_RU = {"income": "Доход", "expense": "Расход"}


def build_operations_xlsx(
    operations: list[dict],
    base_currency: str,
    summary: FinancialSummary | None = None,
    period_label: str = "",
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Операции"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")

    if period_label:
        ws.append([f"FREELABOT — экспорт за {period_label}"])
        ws.merge_cells("A1:J1")
        ws["A1"].font = Font(bold=True, size=12)
        ws.append([])

    ws.append(HEADERS)
    header_row = ws.max_row
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for op in operations:
        ws.append([
            op.get("id"),
            op.get("date", ""),
            TYPE_RU.get(op.get("type", ""), op.get("type", "")),
            op.get("amount_original", 0),
            op.get("currency_original", "RUB"),
            op.get("amount_base", 0),
            op.get("category", ""),
            op.get("counterparty", ""),
            op.get("description", ""),
            op.get("created_at", ""),
        ])

    if summary:
        ws.append([])
        ws.append(["Итого доходы", summary.total_income, base_currency])
        ws.append(["Итого расходы", summary.total_expense, base_currency])
        ws.append(["Прибыль", summary.net, base_currency])
        ws.append(["Налог (расчёт)", summary.tax, base_currency])
        for row in range(ws.max_row - 3, ws.max_row + 1):
            ws.cell(row=row, column=1).font = Font(bold=True)

    for col in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col)
        max_len = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_filename(date_from: str, date_to: str) -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return f"freelabot_{date_from}_{date_to}_{ts}.xlsx"
